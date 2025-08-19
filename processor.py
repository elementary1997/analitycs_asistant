import os
import re
from typing import Tuple, Optional
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference


def _load_input_data(input_file_path: str) -> pd.DataFrame:
    return pd.read_excel(input_file_path, sheet_name="data")


def _clean_dataframe(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, float]:
    columns_to_drop = [
        'Номер', 'Тип операции', 'Сумма', 'Валюта', 'Состояние', 'Номер счета/карты списания'
    ]
    # Be tolerant to missing columns
    existing_to_drop = [c for c in columns_to_drop if c in df.columns]
    df_clean = df.drop(columns=existing_to_drop, errors='ignore')

    words_to_remove = ['N.NOVGOROD', 'RUS', 'Nizhniy Novg', 'NIZJNIY NOVG', 'NIZHNIY NOVG']

    def clean_description(text):
        if pd.isna(text):
            return text
        for word in words_to_remove:
            text = str(text).replace(word, '').strip()
        return text

    if 'Описание' in df_clean.columns:
        df_clean['Описание'] = df_clean['Описание'].apply(clean_description)

    category_sums = df_clean.groupby('Категория')['Сумма в рублях'].sum().reset_index()
    total_expenses_except_invest = df_clean.loc[df_clean['Категория'] != 'На инвестиции', 'Сумма в рублях'].sum()

    return df_clean, category_sums, total_expenses_except_invest


def _write_processed_sheet(writer, sheet_name: str, df_clean: pd.DataFrame,
                          category_sums: pd.DataFrame, total_expenses_except_invest: float) -> None:
    # Write main cleaned table at A1
    df_clean.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)

    ws = writer.sheets[sheet_name]

    # Determine right-side start column for categories table
    right_start_col = df_clean.shape[1] + 2  # leave one gap column

    # Title for categories table
    ws.cell(row=1, column=right_start_col + 1, value='Суммы по категориям')
    ws.cell(row=1, column=right_start_col + 1).font = Font(bold=True)

    # Write categories table at row 2 on the right
    category_sums.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        startrow=1,
        startcol=right_start_col,
    )

    # Colorize categories header row
    header_fill = PatternFill(start_color='FFE5F1FB', end_color='FFE5F1FB', fill_type='solid')
    header_font = Font(bold=True)
    for j in range(category_sums.shape[1]):
        cell = ws.cell(row=2, column=right_start_col + 1 + j)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Add total label below categories table (not under main)
    cat_rows = category_sums.shape[0]
    total_row = 2 + cat_rows + 2  # header at row 2, data starts at 3
    ws.cell(row=total_row, column=right_start_col + 1, value='Итого расходы (без "На инвестиции"):')
    ws.cell(row=total_row, column=right_start_col + 1).font = Font(bold=True)
    # Sum value in the next column to the right of categories "Сумма в рублях"
    try:
        cat_sum_col_idx = list(category_sums.columns).index('Сумма в рублях')
    except ValueError:
        cat_sum_col_idx = category_sums.shape[1] - 1 if category_sums.shape[1] > 0 else 1
    ws.cell(row=total_row, column=right_start_col + 1 + cat_sum_col_idx, value=float(total_expenses_except_invest))

    # Auto-size main table columns
    _autosize_columns(ws, df_clean, start_col=1)
    # Auto-size categories table columns on the right
    _autosize_columns(ws, category_sums, start_col=right_start_col + 1)

    # Format numeric columns for currency – main table
    _format_currency_column(ws, df_clean, start_col=1, header_row=1, target_col_name='Сумма в рублях')
    # Format numeric columns for currency – categories table
    _format_currency_column(ws, category_sums, start_col=right_start_col + 1, header_row=2, target_col_name='Сумма в рублях')

    # Thick borders around tables
    main_end_row = df_clean.shape[0] + 1
    _apply_thick_border(ws, start_row=1, start_col=1, end_row=main_end_row, end_col=df_clean.shape[1])

    if category_sums.shape[0] > 0:
        cat_end_row = 2 + category_sums.shape[0]
        _apply_thick_border(ws, start_row=2, start_col=right_start_col + 1, end_row=cat_end_row, end_col=right_start_col + category_sums.shape[1])

        # Add bar chart under categories table
        try:
            data_ref = Reference(ws, min_col=right_start_col + 1 + cat_sum_col_idx, min_row=2, max_row=cat_end_row)
            cats_ref = Reference(ws, min_col=right_start_col + 1, min_row=3, max_row=cat_end_row)
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.y_axis.title = "Сумма"
            chart.x_axis.title = "Категории"
            # Anchor chart below
            chart_row_anchor = total_row + 2
            anchor_cell = f"{get_column_letter(right_start_col + 1)}{chart_row_anchor}"
            ws.add_chart(chart, anchor_cell)
        except Exception:
            # Skip chart errors silently
            pass


def _autosize_columns(ws, df: pd.DataFrame, start_col: int) -> None:
    for idx, col_name in enumerate(df.columns, start=start_col):
        values = [str(col_name)] + ["" if pd.isna(v) else str(v) for v in df[col_name].tolist()]
        max_len = max(len(v) for v in values) + 2
        ws.column_dimensions[get_column_letter(idx)].width = max(10, min(max_len, 60))


def _format_currency_column(ws, df: pd.DataFrame, start_col: int, header_row: int, target_col_name: str) -> None:
    if target_col_name not in df.columns:
        return
    col_idx = list(df.columns).index(target_col_name) + start_col
    # Data rows start below header
    for r in range(header_row + 1, header_row + 1 + df.shape[0]):
        cell = ws.cell(row=r, column=col_idx)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')


def _apply_thick_border(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    thick = Side(style='thick')
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            top = thick if r == start_row else None
            bottom = thick if r == end_row else None
            left = thick if c == start_col else None
            right = thick if c == end_col else None
            cell.border = Border(
                top=top if top else cell.border.top,
                bottom=bottom if bottom else cell.border.bottom,
                left=left if left else cell.border.left,
                right=right if right else cell.border.right,
            )


def process_report(input_file_path: str, output_file_path: str):
    df = _load_input_data(input_file_path)
    df_clean, category_sums, total_expenses_except_invest = _clean_dataframe(df)

    with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
        # В разовом отчёте всегда один лист — «Очищенные данные»
        _write_processed_sheet(writer, 'Очищенные данные', df_clean, category_sums, total_expenses_except_invest)


_RU_MONTHS = {
    "янв": 1, "январ": 1, "январь": 1,
    "фев": 2, "феврал": 2, "февраль": 2,
    "мар": 3, "март": 3,
    "апр": 4, "апрел": 4, "апрель": 4,
    "май": 5,
    "июн": 6, "июнь": 6,
    "июл": 7, "июль": 7,
    "авг": 8, "август": 8,
    "сен": 9, "сент": 9, "сентябр": 9, "сентябрь": 9,
    "окт": 10, "октябр": 10, "октябрь": 10,
    "ноя": 11, "нояб": 11, "ноябрь": 11,
    "дек": 12, "декабр": 12, "декабрь": 12,
}


# ——— Упрощённая логика без помесячных листов — всё хранится на одном листе ———


def update_master_workbook(master_path: str, input_file_path: str) -> str:
    """Обновляет/создаёт помесячный лист формата MM.YYYY. Возвращает имя листа."""
    df = _load_input_data(input_file_path)
    df_clean, category_sums, total_expenses_except_invest = _clean_dataframe(df)

    sheet_name = _month_year_label(df, input_file_path)

    if not os.path.exists(master_path):
        with pd.ExcelWriter(master_path, engine='xlsxwriter') as writer:
            _write_processed_sheet(writer, sheet_name, df_clean, category_sums, total_expenses_except_invest)
        return sheet_name

    # Прочитать существующие данные помесячного листа, если он есть
    try:
        existing_raw = pd.read_excel(master_path, sheet_name=sheet_name)
        if {'Категория', 'Сумма в рублях'}.issubset(existing_raw.columns):
            existing = existing_raw.dropna(subset=['Категория', 'Сумма в рублях'], how='any')
        else:
            existing = existing_raw
        combined = pd.concat([existing, df_clean], ignore_index=True).drop_duplicates()
    except ValueError:
        combined = df_clean

    # Полная перезапись указанного листа
    from openpyxl import load_workbook
    wb = load_workbook(master_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    wb.create_sheet(title=sheet_name)
    # Sort sheets ascending by MM.YYYY (if all monthly), keep others at end
    try:
        def sort_key(sn: str):
            import re as _re
            m = _re.fullmatch(r"(\d{2})\.(\d{4})", sn)
            if not m:
                return (9999, 99, sn)
            return (int(m.group(2)), int(m.group(1)), sn)
        ordered = sorted(wb.sheetnames, key=sort_key)
        # Reorder by creating a new workbook order via move_sheet
        for name in ordered:
            wb.move_sheet(wb[name], offset=wb.sheetnames.index(name) - wb.sheetnames.index(name))
    except Exception:
        pass
    wb.save(master_path)

    df_clean2, cat2, total2 = _clean_dataframe(combined)
    with pd.ExcelWriter(master_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        _write_processed_sheet(writer, sheet_name, df_clean2, cat2, total2)

    return sheet_name


def _month_year_label(df: pd.DataFrame, filename: str) -> str:
    """Определяет ярлык листа как MM.YYYY из данных; при неудаче — из имени файла."""
    label = _month_year_from_dataframe(df)
    if label:
        return label
    label2 = _month_year_from_filename(filename)
    return label2 or '01.1970'


def _month_year_from_dataframe(df: pd.DataFrame) -> Optional[str]:
    date_cols = [c for c in df.columns if 'дата' in c.lower() or 'date' in c.lower()]
    if not date_cols:
        return None
    col = date_cols[0]

    def parse_date(val: object) -> Optional[pd.Timestamp]:
        if pd.isna(val):
            return None
        s = str(val).strip().lower().replace('\u00a0', ' ').replace(',', ' ')
        # dd.mm.yyyy or dd/mm/yyyy
        m = re.search(r"(\d{1,2})[\.\-/](\d{1,2})[\.\-/](\d{2,4})", s)
        if m:
            day = int(m.group(1)); mon = int(m.group(2)); year = int(m.group(3))
            if year < 100:
                year += 2000
            try:
                return pd.Timestamp(year=year, month=mon, day=day)
            except Exception:
                return None
        # dd mon yy (рус.)
        m2 = re.search(r"(\d{1,2})\s+([а-я\.]+)\s+(\d{2,4})", s)
        if m2:
            mon_raw = m2.group(2).replace('.', '')
            year = int(m2.group(3)); day = int(m2.group(1))
            if year < 100:
                year += 2000
            ru = {
                'янв':1,'январ':1,'январь':1,
                'фев':2,'феврал':2,'февраль':2,
                'мар':3,'март':3,
                'апр':4,'апрел':4,'апрель':4,
                'май':5,
                'июн':6,'июнь':6,
                'июл':7,'июль':7,
                'авг':8,'август':8,
                'сен':9,'сент':9,'сентябр':9,'сентябрь':9,
                'окт':10,'октябр':10,'октябрь':10,
                'ноя':11,'нояб':11,'ноябрь':11,
                'дек':12,'декабр':12,'декабрь':12,
            }
            mon_num = ru.get(mon_raw)
            if mon_num is None:
                for k,v in ru.items():
                    if mon_raw.startswith(k):
                        mon_num = v; break
            if mon_num:
                try:
                    return pd.Timestamp(year=year, month=mon_num, day=day)
                except Exception:
                    return None
        return None

    parsed = df[col].apply(parse_date).dropna()
    if parsed.empty:
        return None
    period = parsed.dt.to_period('M').mode()
    if len(period) == 0:
        return None
    p = period.iloc[0]
    return f"{int(p.month):02d}.{int(p.year):04d}"


def _month_year_from_filename(filename: str) -> Optional[str]:
    base = os.path.basename(filename)
    name, _ = os.path.splitext(base)
    tokens = re.split(r"[\s,_-]+", name)
    # try dd mon yyyy
    s = " ".join(tokens).lower()
    m2 = re.search(r"(\d{1,2})\s+([а-я\.]+)\s+(\d{2,4})", s)
    if m2:
        mon_raw = m2.group(2).replace('.', '')
        year = int(m2.group(3))
        if year < 100:
            year += 2000
        ru = {
            'янв':1,'январ':1,'январь':1,
            'фев':2,'феврал':2,'февраль':2,
            'мар':3,'март':3,
            'апр':4,'апрел':4,'апрель':4,
            'май':5,
            'июн':6,'июнь':6,
            'июл':7,'июль':7,
            'авг':8,'август':8,
            'сен':9,'сент':9,'сентябр':9,'сентябрь':9,
            'окт':10,'октябр':10,'октябрь':10,
            'ноя':11,'нояб':11,'ноябрь':11,
            'дек':12,'декабр':12,'декабрь':12,
        }
        mon_num = ru.get(mon_raw)
        if mon_num is None:
            for k,v in ru.items():
                if mon_raw.startswith(k):
                    mon_num = v; break
        if mon_num:
            return f"{mon_num:02d}.{year:04d}"
    # try dd.mm.yyyy anywhere
    m3 = re.search(r"(\d{1,2})[\.\-/](\d{1,2})[\.\-/](\d{2,4})", s)
    if m3:
        mon = int(m3.group(2)); year = int(m3.group(3))
        if year < 100:
            year += 2000
        return f"{mon:02d}.{year:04d}"
    return None


# Удалена «Общая информация» и вся помесячная логика — по требованию
